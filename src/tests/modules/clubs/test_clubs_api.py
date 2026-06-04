from datetime import datetime, timedelta
from io import BytesIO

import pytest
from openpyxl import load_workbook

from src.modules.bookings.dao import BookingDAO
from src.shared.schemas.schemas import Role


pytestmark = pytest.mark.asyncio


def _xlsx_cell_values(content: bytes) -> list:
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    values = []
    for row in sheet.iter_rows(values_only=True):
        for value in row:
            if value is not None and value != "":
                values.append(value)
    return values


async def test_owner_can_get_club_availability(client, factory):
    city = await factory.create_city()
    owner = await factory.create_user(role=Role.owner, city_id=city.id)
    user = await factory.create_user(city_id=city.id)
    club = await factory.create_club(owner_id=owner.id, city_id=city.id)
    zone = await factory.create_zone(club_id=club.id)
    computer = await factory.create_computer(zone_id=zone.id)
    start_time, end_time = factory.booking_window()

    booking = await factory.create_booking(
        user_id=user.id,
        club_id=club.id,
        zone_id=zone.id,
        computer_id=computer.id,
        start_time=start_time,
        end_time=end_time,
        total_price=1500,
    )
    await BookingDAO.update_check_in_status(booking.id, True)

    response = await client.get(
        f"/clubs/{club.id}/availability",
        params={
            "start_time": start_time.isoformat(),
            "end_time": end_time.isoformat(),
        },
    )

    assert response.status_code == 200
    data = response.json()
    assert len(data) == 1
    assert data[0]["computers"][0]["id"] == computer.id
    assert len(data[0]["computers"][0]["bookings"]) == 1
    assert data[0]["computers"][0]["bookings"][0]["total_price"] == 1500


async def test_owner_can_export_club_statistics_to_excel(client, factory):
    city = await factory.create_city()
    owner = await factory.create_user(role=Role.owner, city_id=city.id, full_name="Owner Name")
    user = await factory.create_user(city_id=city.id)
    club = await factory.create_club(owner_id=owner.id, city_id=city.id, name="Cyber Arena Stats")
    zone = await factory.create_zone(club_id=club.id, name="VIP")
    first_computer = await factory.create_computer(zone_id=zone.id, number=71)
    second_computer = await factory.create_computer(zone_id=zone.id, number=72)
    await factory.create_package(zone_id=zone.id, name="Base hour", duration=1, price=500, is_package=False)
    await factory.create_package(zone_id=zone.id, name="Night package", duration=2, price=900, is_package=True)
    start_time = (datetime.now() - timedelta(days=2)).replace(hour=10, minute=0, second=0, microsecond=0)
    end_time = start_time + timedelta(hours=2)
    first_booking = await factory.create_booking(
        user_id=user.id,
        club_id=club.id,
        zone_id=zone.id,
        computer_id=first_computer.id,
        start_time=start_time,
        end_time=end_time,
        total_price=900,
    )
    second_booking = await factory.create_booking(
        user_id=user.id,
        club_id=club.id,
        zone_id=zone.id,
        computer_id=first_computer.id,
        start_time=end_time + timedelta(hours=1),
        end_time=end_time + timedelta(hours=2),
        total_price=500,
    )
    third_booking = await factory.create_booking(
        user_id=user.id,
        club_id=club.id,
        zone_id=zone.id,
        computer_id=second_computer.id,
        start_time=end_time + timedelta(hours=3),
        end_time=end_time + timedelta(hours=4),
        total_price=500,
    )
    await BookingDAO.update_check_in_status(first_booking.id, True)
    await BookingDAO.update_check_in_status(second_booking.id, True)
    await BookingDAO.update_check_in_status(third_booking.id, True)

    response = await client.post(
        f"/clubs/{club.id}/statistics/export",
        headers=factory.auth_headers(factory.access_token(owner.id)),
        json={
            "start_date": start_time.date().isoformat(),
            "end_date": start_time.date().isoformat(),
        },
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment;" in response.headers["content-disposition"]
    assert response.headers["content-disposition"].endswith('.xlsx"')

    values = _xlsx_cell_values(response.content)
    assert "Cyber Arena Stats" in values
    assert "Owner Name" in values
    assert "Payment type price" in values
    assert "VIP" in values
    assert 71 in values
    assert 72 in values
    assert "Zone total amount" in values
    assert "Total amount" in values
    assert "Night package" in values
    assert "Base hour" in values
    assert 900 in values
    assert 500 in values
    assert 1900 in values
    assert values.count(1400) == 1
    assert values.count(1900) == 2
